"""
exporter.py — Exportação dos lançamentos para Excel (.xlsx) com formatação profissional.

Gera três abas:
  - Extrato:          lançamentos contábeis
  - Resumo:           totais e saldo do período
  - Linhas Ignoradas: relatório de linhas que não puderam ser processadas

Suporta dois modos de saída:
  - Disco (CLI):    export_to_excel(result, path)   → salva arquivo em disco
  - Memória (Web):  export_to_bytes(result)          → retorna BytesIO para Streamlit
"""

from io import BytesIO
from pathlib import Path
from typing import List, Optional, Union

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import Lancamento, ParseResult

# ── Constantes de estilo ──────────────────────────────────────────────────────

_HEADER_FILL   = PatternFill("solid", start_color="1F4E79")   # Azul escuro
_HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_DATA_FONT     = Font(name="Arial", size=10)
_ALT_FILL      = PatternFill("solid", start_color="DCE6F1")   # Azul claro alternado
_DEBIT_FONT    = Font(name="Arial", color="C00000", size=10)  # Vermelho para débitos
_CREDIT_FONT   = Font(name="Arial", color="375623", size=10)  # Verde para créditos
_BORDER_THIN   = Side(style="thin", color="BFBFBF")
_CELL_BORDER   = Border(
    left=_BORDER_THIN, right=_BORDER_THIN,
    top=_BORDER_THIN,  bottom=_BORDER_THIN,
)
_BR_CURRENCY   = '#,##0.00'   # Formato numérico: 1.234,56 (depende do locale do Excel)

_COLUMNS = ["Data", "Descrição / Histórico", "Débito (R$)", "Crédito (R$)"]
_COL_WIDTHS = [14, 55, 18, 18]


# ── Exportação principal ──────────────────────────────────────────────────────

def _build_workbook(result: Union[ParseResult, List[Lancamento]]) -> "Workbook":
    """Constrói e retorna o workbook openpyxl formatado (sem salvar em disco)."""
    if isinstance(result, list):
        lancamentos, ignoradas = result, []
    else:
        lancamentos, ignoradas = result.lancamentos, result.ignoradas

    records = [
        {
            "Data":                  l.data,
            "Descrição / Histórico": l.descricao,
            "Débito (R$)":           l.debito,
            "Crédito (R$)":          l.credito,
        }
        for l in lancamentos
    ]
    df = pd.DataFrame(records, columns=_COLUMNS)

    # Salva em buffer temporário para o openpyxl conseguir abrir
    buf = BytesIO()
    df.to_excel(buf, index=False, sheet_name="Extrato", engine="openpyxl")
    buf.seek(0)

    wb = load_workbook(buf)
    ws = wb.active
    ws.title = "Extrato"

    _apply_formatting(ws, len(df))
    _add_summary_sheet(wb, df)
    _add_ignored_sheet(wb, ignoradas)

    return wb


def export_to_excel(result: Union[ParseResult, List[Lancamento]], output_path: Path) -> None:
    """
    Modo CLI: constrói o workbook e salva em disco.

    Args:
        result:      ParseResult ou lista de Lancamento.
        output_path: Caminho de destino (.xlsx).
    """
    wb = _build_workbook(result)
    wb.save(str(output_path))


def export_to_bytes(result: Union[ParseResult, List[Lancamento]]) -> BytesIO:
    """
    Modo Web: constrói o workbook e retorna um BytesIO pronto para
    ser entregue via st.download_button() do Streamlit.

    O buffer já está posicionado em seek(0) e pode ser lido diretamente.

    Args:
        result: ParseResult ou lista de Lancamento.

    Returns:
        BytesIO com o conteúdo do arquivo .xlsx.
    """
    wb = _build_workbook(result)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _apply_formatting(ws, num_rows: int) -> None:
    """Aplica estilos à aba principal."""
    # Cabeçalho
    for col_idx, (col_name, width) in enumerate(zip(_COLUMNS, _COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _CELL_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"  # Congela o cabeçalho

    # Linhas de dados
    for row_idx in range(2, num_rows + 2):
        is_odd = (row_idx % 2 == 0)

        for col_idx in range(1, 5):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _CELL_BORDER

            if is_odd:
                cell.fill = _ALT_FILL

            if col_idx == 1:  # Data
                cell.font = _DATA_FONT
                cell.alignment = Alignment(horizontal="center")

            elif col_idx == 2:  # Descrição
                cell.font = _DATA_FONT
                cell.alignment = Alignment(horizontal="left", wrap_text=False)

            elif col_idx == 3:  # Débito
                cell.font = _DEBIT_FONT
                cell.alignment = Alignment(horizontal="right")
                if cell.value is not None:
                    cell.number_format = _BR_CURRENCY

            elif col_idx == 4:  # Crédito
                cell.font = _CREDIT_FONT
                cell.alignment = Alignment(horizontal="right")
                if cell.value is not None:
                    cell.number_format = _BR_CURRENCY

    ws.row_dimensions[1].height = 28

    # Linha de totais
    total_row = num_rows + 2
    ws.cell(total_row, 1).value = "TOTAL"
    ws.cell(total_row, 1).font = Font(name="Arial", bold=True, size=10)
    ws.cell(total_row, 1).alignment = Alignment(horizontal="center")
    ws.cell(total_row, 1).border = _CELL_BORDER

    ws.cell(total_row, 2).value = f"Total de lançamentos: {num_rows}"
    ws.cell(total_row, 2).font = Font(name="Arial", bold=True, size=10)
    ws.cell(total_row, 2).border = _CELL_BORDER

    for col_idx in (3, 4):
        col_letter = get_column_letter(col_idx)
        formula_cell = ws.cell(total_row, col_idx)
        formula_cell.value = f"=SUM({col_letter}2:{col_letter}{num_rows + 1})"
        formula_cell.number_format = _BR_CURRENCY
        formula_cell.font = Font(name="Arial", bold=True, size=10,
                                 color=("C00000" if col_idx == 3 else "375623"))
        formula_cell.alignment = Alignment(horizontal="right")
        formula_cell.border = _CELL_BORDER
        formula_cell.fill = PatternFill("solid", start_color="D9D9D9")


def _add_summary_sheet(wb, df: pd.DataFrame) -> None:
    """Cria uma aba 'Resumo' com totais e saldo calculados por fórmula."""
    ws = wb.create_sheet("Resumo")

    total_deb = df["Débito (R$)"].sum()
    total_cred = df["Crédito (R$)"].sum()

    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    data_font   = Font(name="Arial", size=10)

    rows = [
        ("RESUMO DO EXTRATO", None),
        ("Total de Lançamentos", len(df)),
        ("Total de Débitos (R$)", "=Extrato!C{}".format(len(df) + 2)),
        ("Total de Créditos (R$)", "=Extrato!D{}".format(len(df) + 2)),
        ("Saldo do Período (R$)", "=B5-B4"),
    ]

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    for r_idx, (label, value) in enumerate(rows, start=1):
        cell_a = ws.cell(r_idx, 1, label)
        cell_b = ws.cell(r_idx, 2, value)

        if r_idx == 1:
            cell_a.font = header_font
            cell_a.fill = header_fill
            cell_b.fill = header_fill
        else:
            cell_a.font = Font(name="Arial", bold=True, size=10)
            cell_b.font = data_font
            cell_b.alignment = Alignment(horizontal="right")

        if r_idx in (3, 4, 5):
            cell_b.number_format = _BR_CURRENCY
            if r_idx == 5:  # Saldo
                cell_b.font = Font(name="Arial", bold=True, size=10,
                                   color=("375623" if total_cred >= total_deb else "C00000"))

        cell_a.border = _CELL_BORDER
        cell_b.border = _CELL_BORDER


def _add_ignored_sheet(wb, ignoradas) -> None:
    """
    Cria a aba 'Linhas Ignoradas' com o relatório detalhado de linhas não processadas.
    Se não houver linhas ignoradas, cria a aba vazia com mensagem de sucesso.
    """
    ws = wb.create_sheet("Linhas Ignoradas")

    # Paleta de cores
    HDR_FILL  = PatternFill("solid", start_color="7B2D2D")   # Vermelho escuro
    HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    OK_FILL   = PatternFill("solid", start_color="E2EFDA")   # Verde claro
    OK_FONT   = Font(name="Arial", bold=True, color="375623", size=11)
    DATA_FONT = Font(name="Arial", size=9)
    ALT_FILL  = PatternFill("solid", start_color="FCE4E4")   # Vermelho muito claro
    NUM_FONT  = Font(name="Arial", size=9, color="7B2D2D", bold=True)

    if not ignoradas:
        # Sem erros — exibe mensagem de sucesso
        ws.merge_cells("A1:E1")
        cell = ws["A1"]
        cell.value = "✓  Nenhuma linha foi ignorada — todas as transações detectadas foram processadas com sucesso."
        cell.font = OK_FONT
        cell.fill = OK_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        ws.column_dimensions["A"].width = 90
        return

    # Cabeçalho da tabela
    headers = ["Nº Linha", "Pág. Estimada", "Motivo", "Conteúdo da Linha"]
    col_widths = [10, 14, 48, 80]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(1, col_idx, header)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _CELL_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    # Agrupa motivos para colorir linhas por categoria de erro
    motivos_unicos = list(dict.fromkeys(i.motivo for i in ignoradas))
    cor_por_motivo = {}
    cores_alternadas = ["FCE4E4", "FFF2CC", "DEEBF7", "EAF0FB"]
    for idx, motivo in enumerate(motivos_unicos):
        cor_por_motivo[motivo] = cores_alternadas[idx % len(cores_alternadas)]

    # Linhas de dados
    for row_idx, ign in enumerate(ignoradas, start=2):
        fill_color = cor_por_motivo.get(ign.motivo, "FCE4E4")
        row_fill = PatternFill("solid", start_color=fill_color)

        valores = [ign.numero_linha, ign.pagina_estimada, ign.motivo, ign.conteudo_resumido(78)]
        aligns  = ["center", "center", "left", "left"]
        fonts   = [NUM_FONT, DATA_FONT, DATA_FONT, DATA_FONT]

        for col_idx, (val, align, font) in enumerate(zip(valores, aligns, fonts), start=1):
            cell = ws.cell(row_idx, col_idx, val)
            cell.font = font
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=(col_idx == 4))
            cell.border = _CELL_BORDER

        ws.row_dimensions[row_idx].height = 18

    # Rodapé com contagem por motivo
    total_row = len(ignoradas) + 2
    ws.cell(total_row + 1, 1).value = "RESUMO POR MOTIVO"
    ws.cell(total_row + 1, 1).font = Font(name="Arial", bold=True, size=10)
    ws.merge_cells(f"A{total_row + 1}:D{total_row + 1}")

    motivos_ordenados = sorted(
        motivos_unicos,
        key=lambda m: sum(1 for i in ignoradas if i.motivo == m),
        reverse=True,
    )
    for offset, motivo in enumerate(motivos_ordenados, start=total_row + 2):
        count_val = sum(1 for i in ignoradas if i.motivo == motivo)
        ws.cell(offset, 1).value = count_val
        ws.cell(offset, 1).font = NUM_FONT
        ws.cell(offset, 1).alignment = Alignment(horizontal="center")

        ws.merge_cells(f"B{offset}:D{offset}")
        ws.cell(offset, 2).value = motivo
        ws.cell(offset, 2).font = Font(name="Arial", size=9)
        ws.cell(offset, 2).fill = PatternFill("solid", start_color=cor_por_motivo[motivo])
